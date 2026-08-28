"""业务编排：建活动、投稿、文件落盘、导出、简报、智能问答。"""
import csv
import io
import json
import secrets
import uuid
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from fastapi import HTTPException, UploadFile
from PIL import Image
from sqlalchemy.orm import Session

from db import SessionLocal
from config import settings
from models import Campaign, Applicant, Work, CheckReport
from schemas import CampaignCreate, CheckReportOut
from checks import run_check
from tools import campaign_overview, list_works as list_works_tool, run_sql as run_sql_tool
from agent import llm_brief, plan_query, synthesize_answer


def create_campaign(db: Session, data: CampaignCreate) -> Campaign:
    camp = Campaign(
        title=data.title,
        description=data.description,
        deadline=data.deadline,
        image_formats=data.image_formats or settings.default_image_formats,
        max_image_mb=data.max_image_mb or settings.default_max_image_mb,
        link_token=secrets.token_urlsafe(16),
    )
    db.add(camp)
    db.commit()
    db.refresh(camp)
    return camp


def persist_file(upload: UploadFile | None, allowed_exts: list, max_mb: float,
                 must_be_image: bool = False):
    """保存上传文件。返回 (stored_name, error, parse_error, size_mb)。"""
    allowed = [e.strip().lower().lstrip(".") for e in allowed_exts]
    if upload is None or upload.filename is None:
        return None, "未选择文件", None, None

    ext = Path(upload.filename).suffix.lstrip(".").lower()
    if ext not in allowed:
        return None, f"文件格式 .{ext} 不在允许范围（{'、'.join(allowed)}）", None, None

    data = upload.file.read()
    if not data:
        return None, "文件内容为空", None, None
    size_mb = len(data) / (1024 * 1024)
    if size_mb > max_mb:
        return None, f"文件大小 {size_mb:.1f}MB 超过上限 {max_mb}MB", None, None

    parse_error = None
    if must_be_image:
        try:
            img = Image.open(io.BytesIO(data))
            img.verify()
        except Exception as e:  # noqa: BLE001 图片损坏 -> 记录"无法解析"，不判缺失
            parse_error = str(e)

    stored = f"{uuid.uuid4().hex}.{ext}"
    (settings.upload_dir / stored).write_bytes(data)
    return stored, None, parse_error, size_mb


def submit_submission(db: Session, campaign: Campaign, name: str, phone: str, email: str,
                      wechat: str, resume_file: UploadFile | None, works_data: list,
                      image_files: list) -> tuple[Applicant, CheckReportOut]:
    if not (name or "").strip():
        raise HTTPException(400, "姓名不能为空")

    # 简历（pdf/docx）
    resume_name = None
    if resume_file is not None:
        resume_name, err, _, _ = persist_file(resume_file, settings.resume_formats,
                                              settings.resume_max_mb, must_be_image=False)
        if err:
            raise HTTPException(400, f"简历：{err}")

    applicant = Applicant(
        campaign_id=campaign.id,
        name=name.strip(),
        phone=phone or "",
        email=email or "",
        wechat=wechat or "",
        resume_path=resume_name or "",
        status="submitted",
    )
    db.add(applicant)
    db.flush()

    works = list(works_data or [])
    works_for_check: list = []
    for i, wd in enumerate(works):
        img = image_files[i] if i < len(image_files) else None
        stored = ""
        parse_err = None
        size_mb = None
        if img is not None and img.filename:
            stored, err, parse_err, size_mb = persist_file(
                img, campaign.image_formats.split(","), campaign.max_image_mb, must_be_image=True)
            if err:
                raise HTTPException(400, f"作品{i + 1}：{err}")

        w = Work(
            applicant_id=applicant.id,
            title=(wd.get("title") or "").strip(),
            dimensions=(wd.get("dimensions") or "").strip(),
            medium=(wd.get("medium") or "").strip(),
            school=(wd.get("school") or "").strip(),
            price=(wd.get("price") or "").strip(),
            image_path=stored,
        )
        db.add(w)
        works_for_check.append({
            "title": w.title, "dimensions": w.dimensions, "medium": w.medium,
            "school": w.school, "price": w.price, "image_path": stored,
            "image_size_mb": size_mb, "image_parse_error": parse_err,
        })

    db.flush()

    report = run_check(
        {"name": applicant.name, "phone": applicant.phone, "email": applicant.email,
         "wechat": applicant.wechat, "resume_path": applicant.resume_path},
        works_for_check,
        {"image_formats": campaign.image_formats, "max_image_mb": campaign.max_image_mb},
    )

    cr = CheckReport(
        applicant_id=applicant.id,
        missing=json.dumps(report.missing, ensure_ascii=False),
        format_issues=json.dumps(report.format_issues, ensure_ascii=False),
        notes=report.notes,
    )
    db.add(cr)
    applicant.status = "checked"
    db.commit()
    db.refresh(applicant)
    return applicant, report


def build_overview(campaign_id: int) -> dict:
    return campaign_overview(campaign_id)


def export_rows(campaign_id: int, medium: str | None = None, school: str | None = None) -> list:
    with SessionLocal() as db:
        applicants = db.query(Applicant).filter(Applicant.campaign_id == campaign_id).all()
        rows = []
        for a in applicants:
            works = a.works
            if medium:
                works = [w for w in works if w.medium == medium]
            if school:
                works = [w for w in works if w.school == school]
            if not works:
                rows.append([a.name, a.phone, a.email, a.wechat, "", "", "", "", "", a.status])
            for w in works:
                rows.append([a.name, a.phone, a.email, a.wechat, w.title, w.dimensions,
                             w.medium, w.school, w.price, a.status])
        return rows


def deterministic_brief(overview: dict) -> str:
    lines = [f"# {overview['campaign_title']} · 工作简报", ""]
    lines.append(f"**艺术家信息**：共 {overview['applicant_count']} 位投稿艺术家。")
    if overview["artist_list"]:
        lines.append("艺术家名单：" + "、".join(a["name"] for a in overview["artist_list"]))
    lines.append("")
    lines.append(f"**作品数量**：共 {overview['work_count']} 件作品。")
    lines.append("")
    lines.append("**学校分布**：" + (
        "、".join(f"{k} {v}件" for k, v in overview["school_distribution"].items())
        if overview["school_distribution"] else "暂无"))
    lines.append("")
    lines.append("**作品种类**：" + (
        "、".join(f"{k} {v}件" for k, v in overview["medium_distribution"].items())
        if overview["medium_distribution"] else "暂无"))
    return "\n".join(lines)


def export_zip(campaign_id: int, medium: str | None = None, school: str | None = None) -> bytes:
    """导出 ZIP：export.csv + images/ + resumes/，图片与简历随表格一起导出。"""
    with SessionLocal() as db:
        if not db.get(Campaign, campaign_id):
            raise ValueError("活动不存在")
        applicants = db.query(Applicant).filter(Applicant.campaign_id == campaign_id).all()

        csv_buf = io.StringIO()
        writer = csv.writer(csv_buf)
        writer.writerow(["姓名", "电话", "邮箱", "微信", "作品名", "尺寸", "画种", "毕业院校",
                         "价格", "照片文件", "简历文件", "投稿状态"])
        files: list = []

        for a in applicants:
            works = a.works
            if medium:
                works = [w for w in works if w.medium == medium]
            if school:
                works = [w for w in works if w.school == school]

            resume_arc = ""
            if a.resume_path:
                rp = settings.upload_dir / a.resume_path
                if rp.exists():
                    resume_arc = f"resumes/{a.id}_{Path(a.resume_path).name}"
                    files.append((resume_arc, rp))

            if not works:
                writer.writerow([a.name, a.phone, a.email, a.wechat, "", "", "", "", "",
                                 "", resume_arc, a.status])
            for w in works:
                img_arc = ""
                if w.image_path:
                    ip = settings.upload_dir / w.image_path
                    if ip.exists():
                        img_arc = f"images/{w.id}_{Path(w.image_path).name}"
                        files.append((img_arc, ip))
                writer.writerow([a.name, a.phone, a.email, a.wechat, w.title, w.dimensions,
                                 w.medium, w.school, w.price, img_arc, resume_arc, a.status])

        csv_bytes = csv_buf.getvalue().encode("utf-8-sig")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("export.csv", csv_bytes)
        for arc, full in files:
            zf.write(full, arc)
    return zip_buf.getvalue()


EXPORT_COLUMNS = [
    ("name", "姓名"), ("phone", "电话"), ("email", "邮箱"), ("wechat", "微信"),
    ("title", "作品名"), ("dimensions", "尺寸"), ("medium", "画种"), ("school", "毕业院校"),
    ("price", "价格"), ("image", "照片"), ("resume", "简历"), ("status", "投稿状态"),
]


def export_xlsx(campaign_id: int, columns: list | None = None, medium: str | None = None,
                school: str | None = None, has_resume: str | None = None) -> bytes:
    """导出 Excel：图片内嵌到单元格，列可筛选。"""
    col_map = dict(EXPORT_COLUMNS)
    selected = [c for c in (columns or []) if c in col_map] or [c for c, _ in EXPORT_COLUMNS]

    with SessionLocal() as db:
        if not db.get(Campaign, campaign_id):
            raise ValueError("活动不存在")
        applicants = db.query(Applicant).filter(Applicant.campaign_id == campaign_id).all()
        if has_resume == "yes":
            applicants = [a for a in applicants if a.resume_path]
        elif has_resume == "no":
            applicants = [a for a in applicants if not a.resume_path]

        rows: list = []
        for a in applicants:
            works = a.works
            if medium:
                works = [w for w in works if w.medium == medium]
            if school:
                works = [w for w in works if w.school == school]
            if not works:
                rows.append((a, None))
            for w in works:
                rows.append((a, w))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "投稿明细"
    ws.append([col_map[c] for c in selected])

    image_col_idx = selected.index("image") + 1 if "image" in selected else None
    for a, w in rows:
        values = {
            "name": a.name, "phone": a.phone, "email": a.email, "wechat": a.wechat,
            "title": w.title if w else "", "dimensions": w.dimensions if w else "",
            "medium": w.medium if w else "", "school": w.school if w else "",
            "price": w.price if w else "", "image": "",
            "resume": Path(a.resume_path).name if a.resume_path else "",
            "status": a.status,
        }
        ws.append([values[c] for c in selected])
        if image_col_idx and w and w.image_path:
            ip = settings.upload_dir / w.image_path
            if ip.exists():
                r = ws.max_row
                img = XLImage(str(ip))
                img.width = 120
                img.height = 90
                ws.add_image(img, f"{get_column_letter(image_col_idx)}{r}")
                ws.row_dimensions[r].height = 70

    for idx, key in enumerate(selected, 1):
        letter = get_column_letter(idx)
        widths = {"image": 18, "email": 22, "title": 16, "school": 16}
        ws.column_dimensions[letter].width = widths.get(key, 12)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_brief(campaign_id: int) -> dict:
    overview = campaign_overview(campaign_id)
    brief = llm_brief(overview) if settings.model_api_key else None
    if brief:
        return {"overview": overview, "brief": brief, "source": "llm"}
    return {"overview": overview, "brief": deterministic_brief(overview), "source": "deterministic"}


def answer_question(question: str, campaign_id: int) -> dict:
    if not settings.model_api_key:
        return {"answer": "未配置模型 API Key，无法智能问答。当前可查看：活动总览、作品明细筛选、只读 SQL 查询。",
                "tool": None, "used_llm": False}
    plan = plan_query(question, campaign_id)
    if not plan:
        return {"answer": "智能问答暂不可用，请稍后重试或使用固定统计功能。", "tool": None, "used_llm": False}
    tool = (plan.get("tool") or "").strip()
    args = plan.get("args") or {}
    try:
        if tool == "overview":
            result = campaign_overview(campaign_id)
        elif tool == "list_works":
            result = list_works_tool(campaign_id, medium=args.get("medium"), school=args.get("school"))
        elif tool == "run_sql":
            result = run_sql_tool(args.get("sql") or "")
        else:
            return {"answer": "无法识别的查询类型。", "tool": tool, "used_llm": False}
    except Exception as e:  # noqa: BLE001 工具执行失败 -> 返回可读错误
        return {"answer": f"查询执行失败：{e}", "tool": tool, "used_llm": False}

    ans = synthesize_answer(question, result)
    if ans is None:
        ans = json.dumps(result, ensure_ascii=False, default=str)
    return {"answer": ans, "tool": tool, "used_llm": True, "result": result}
